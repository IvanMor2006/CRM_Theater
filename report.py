from datetime import datetime, date
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class Report:
    def __init__(self, parent, grid, title):
        self.parent = parent
        self.grid = grid
        self.title = title

    def export(self):
        if not self.grid.rows:
            messagebox.showinfo('Предупреждение', 'Отсутствуют данные для экспорта!', icon='warning', parent=self.parent)
            return False
        
        filename = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx'), ('All files', '*.*')],
            initialfile=f'{self.title}_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx',
            parent=self.parent
        )

        if not filename:
            return False
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = self.title[:31]

            header_font = Font(bold=True, size=11, color='ffffff')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            headers = [field[0] for field in self.grid.FIELDS]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            for row_i, row_data in enumerate(self.grid.rows, 2):
                for col_i, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_i, column=col_i, value=value)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center')

                    if isinstance(value, datetime):
                        cell.number_format = 'YYYY-MM-DD HH:MM:SS'
                    elif isinstance(value, date):
                        cell.number_format = 'YYYY-MM-DD'
                    elif isinstance(value, (int, float)):
                        if col_i - 1 < len(headers) and 'Цена' in headers[col_i - 1]:
                            cell.number_format = '#,##0.00'
            
            for col in range(1, len(headers) + 1):
                max_length = 0
                column_letter = get_column_letter(col)
                for row in range(1, len(self.grid.rows) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            ws.print_title_rows = '1:1'
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            
            wb.save(filename)
            messagebox.showinfo('Успех', f'Отчет успешно сохранен в файл:\n{filename}', icon='info', parent=self.parent)
            return True
            
        except Exception as e:
            messagebox.showerror('Ошибка', f'Ошибка при сохранении файла:\n{str(e)}', icon='error', parent=self.parent)
            return False